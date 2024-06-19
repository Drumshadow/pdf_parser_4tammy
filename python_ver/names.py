import pandas as pd
import pdfquery
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from datetime import date



def resident_cc_total(xml_file):
    resident_names = []
    code_total = []
    case_total = []
    
    try:
        # Parse the XML file
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Function to recursively search for the string in all elements
        def search_element(element):
            for subelement in element:
                if subelement.text and "Resident:" in subelement.text:
                    start_index = subelement.text.find("Resident:")
                    resident_names.append(subelement.text[start_index+9:].strip())

                if subelement.text and "Code Total:" in subelement.text:
                    start_index = subelement.text.find("Code Total:")
                    code_total.append(subelement.text[start_index+11:].strip())

                if subelement.text and "Case Total:" in subelement.text:
                    start_index = subelement.text.find("Case Total:")
                    case_total.append(subelement.text[start_index+11:].strip())
                    
                search_element(subelement)
        # Start searching from the root element
        search_element(root)
    
    except ET.ParseError as e:
        print(f"Error parsing XML file: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

    del resident_names[1::2]
    del code_total[1::2]
    del case_total[1::2]

    return resident_names, code_total, case_total

def numOfDays(date1, date2):
  #check which date is greater to avoid days output in -ve number
    if date2 > date1:   
        return (date2-date1).days
    else:
        return (date1-date2).days

def find_case_created(xml_file):
    case_date = []
    date_entered = []
    
    try:
        # Parse the XML file
        tree = ET.parse(xml_file)
        root = tree.getroot()

        # Function to recursively search for the string in all elements
        def search_element(element):
            for subelement in element:
                if subelement.text and "Case Date:" in subelement.text:
                    # Capture the text starting from "Case created:" to the end
                    start_index = subelement.text.find("Case Date:")
                    case_date.append(subelement.text[start_index+11:].strip())
                
                if subelement.text and "Date Entered:" in subelement.text:
                    # Capture the text starting from "Case created:" to the end
                    start_index = subelement.text.find("Date Entered:")
                    date_entered.append(subelement.text[start_index+14:].strip())
                search_element(subelement)
        # Start searching from the root element
        search_element(root)
    
    except ET.ParseError as e:
        print(f"Error parsing XML file: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

    date_dif = []
    for x in range(len(case_date)):
        s1 = date_entered[x].split("/")
        s2 = case_date[x].split("/")
        date1 = date(int(s1[2]), int(s1[0]), int(s1[1]))
        date2 = date(int(s2[2]), int(s2[0]), int(s2[1]))
        date_dif.append(numOfDays(date1, date2))

    return date_dif

def write_to_excel(names, case_total, avg, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Names Codes Cases"
    
    # Write header
    ws.append(["Names", "Case Total", "Average"])
    
    # Write data
    for entry in names:
        ws.append([entry])

    for row_index, entry in enumerate(case_total, start=2):
        ws.cell(row=row_index, column=2, value=entry)
    
    for row_index, entry in enumerate(avg, start=2):
        ws.cell(row=row_index, column=3, value=entry)

    # Save the workbook
    wb.save(output_file)
    print(f"Data successfully written to {output_file}")

def calc_avgs(names, code_total, date_difs):
    sum = 0
    ndx = 0
    avgs = []
    y = 0
    for x in range(len(names)):
        while y < int(code_total[x]):
            sum+=date_difs[ndx]
            ndx+=1
            y+=1
        y = 0
        avgs.append((sum/int(code_total[x])))
        sum = 0
    return avgs

def process_files(output_excel, xml_out):

    # Find cases created
    names, code_total, case_total = resident_cc_total(xml_out)
    date_diffs = find_case_created(xml_out)
    avg = calc_avgs(names, code_total, date_diffs)
    # Write to Excel
    write_to_excel(names, case_total, avg, output_excel)

process_files('out2.xlsx', 'out2.xml')